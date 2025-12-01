"""
파일 처리 및 텍스트 추출 유틸리티 함수
"""
import os
from typing import List, Tuple
from pypdf import PdfReader
from docx import Document
import openpyxl


def extract_text_from_pdf(file) -> str:
    """PDF 파일에서 텍스트 추출"""
    try:
        pdf_reader = PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"[PDF 추출 오류: {str(e)}]"


def extract_text_from_docx(file) -> str:
    """DOCX 파일에서 텍스트 추출"""
    try:
        doc = Document(file)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    except Exception as e:
        return f"[DOCX 추출 오류: {str(e)}]"


def extract_text_from_xlsx(file) -> str:
    """XLSX 파일에서 텍스트 추출"""
    try:
        workbook = openpyxl.load_workbook(file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n=== Sheet: {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                text += row_text + "\n"
        return text
    except Exception as e:
        return f"[XLSX 추출 오류: {str(e)}]"


def extract_text_from_txt(file) -> str:
    """TXT 파일에서 텍스트 추출"""
    try:
        content = file.read()
        if isinstance(content, bytes):
            return content.decode('utf-8', errors='ignore')
        return content
    except Exception as e:
        return f"[TXT 추출 오류: {str(e)}]"


def extract_text_from_file(file, filename: str) -> str:
    """
    파일 확장자에 따라 적절한 추출 함수 호출
    
    Args:
        file: 업로드된 파일 객체
        filename: 파일명
        
    Returns:
        추출된 텍스트
    """
    ext = os.path.splitext(filename)[1].lower()
    
    if ext == '.pdf':
        return extract_text_from_pdf(file)
    elif ext in ['.docx', '.doc']:
        return extract_text_from_docx(file)
    elif ext in ['.xlsx', '.xls']:
        return extract_text_from_xlsx(file)
    elif ext == '.txt':
        return extract_text_from_txt(file)
    else:
        return f"[지원하지 않는 파일 형식: {ext}]"


def chunk_text(text: str, chunk_size: int = 1000, chunk_overlap: int = 200) -> List[str]:
    """
    텍스트를 청크로 분할
    
    Args:
        text: 원본 텍스트
        chunk_size: 청크 크기
        chunk_overlap: 청크 간 중첩 크기
        
    Returns:
        청크 리스트
    """
    chunks = []
    start = 0
    text_length = len(text)
    
    while start < text_length:
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk)
        start += (chunk_size - chunk_overlap)
    
    return chunks


def format_documents_for_prompt(documents: List[Tuple[str, str, str]]) -> str:
    """
    문서 정보를 프롬프트용 텍스트로 포맷팅
    
    Args:
        documents: [(카테고리, 파일명, 텍스트), ...] 형태의 리스트
        
    Returns:
        포맷팅된 텍스트
    """
    formatted = ""
    for category, filename, text in documents:
        formatted += f"\n{'='*50}\n"
        formatted += f"📁 카테고리: {category}\n"
        formatted += f"📄 파일명: {filename}\n"
        formatted += f"{'='*50}\n"
        formatted += f"{text[:2000]}...\n"  # 각 문서당 최대 2000자
    
    return formatted

